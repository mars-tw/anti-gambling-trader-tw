"""通用交易資料載入器。

支援 CSV / JSON / Excel,並且能「自動辨識欄位」 — 因為每家券商、
每個交易所匯出的欄位名稱都不一樣(中文、英文、大小寫、底線……)。
這裡用同義詞對照表盡量自動對應,對應不到的才要求使用者明確指定。
"""

from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

from ..models import Market, Side, Trade, TradeLog
from .costs import estimate_round_trip_cost

# ── 欄位同義詞對照(全部轉小寫、去空白後比對)──────────────────
# 涵蓋台股券商、美股券商、加密交易所的常見匯出欄位。
FIELD_SYNONYMS: dict[str, list[str]] = {
    "symbol": [
        "symbol", "ticker", "代號", "股票代號", "標的", "商品", "幣別", "pair",
        "instrument", "code", "stockcode", "證券代號",
        # 台股券商 / 幣安 / IBKR
        "證券名稱", "股票名稱", "商品名稱", "market", "underlyingsymbol", "contract",
    ],
    "side": [
        "side", "方向", "買賣", "買賣別", "交易別", "buysell", "direction", "type",
        "action", "long_short", "多空",
    ],
    "entry_time": [
        "entry_time", "entrytime", "進場時間", "買進時間", "open_time", "opentime",
        "成交日期", "進場日", "date", "buy_date", "開倉時間",
        # 台股 / 幣安 / IBKR
        "成交日", "交易日期", "委託日期", "date(utc)", "tradedate", "datetime",
    ],
    "exit_time": [
        "exit_time", "exittime", "出場時間", "賣出時間", "close_time", "closetime",
        "平倉時間", "sell_date", "exit_date", "平倉日",
    ],
    "entry_price": [
        "entry_price", "entryprice", "進場價", "買進價", "open_price", "成交價",
        "買價", "cost", "成本", "開倉價", "buy_price",
        # 台股 / 幣安 / IBKR(注意:成交價金/價金是「總額」非單價,不納入)
        "成交均價", "price", "tradeprice", "成交單價",
    ],
    "exit_price": [
        "exit_price", "exitprice", "出場價", "賣出價", "close_price", "平倉價",
        "賣價", "sell_price",
    ],
    "quantity": [
        "quantity", "qty", "數量", "股數", "張數", "張", "size", "amount", "volume",
        "口數", "成交數量", "成交股數", "shares", "成交量",
    ],
    "fees": [
        "fees", "fee", "手續費", "費用", "成本費用", "commission", "手續費及稅",
        "交易成本", "total_fee",
        # 台股稅費分項 / 幣安 / IBKR
        "證交稅", "交易稅", "手續費及交易稅", "ibcommission",
    ],
    "pnl": [
        "pnl", "profit", "損益", "盈虧", "已實現損益", "realized_pnl", "獲利",
        "net_pnl", "賺賠", "return",
        # 台股 / 幣安 / IBKR
        "損益金額", "淨收付", "淨收付金額", "realized profit", "fifopnlrealized",
    ],
    "tag": [
        "tag", "策略", "strategy", "標籤", "備註", "note", "remark", "策略名稱",
        "setup", "進場理由",
    ],
}

# side 欄位的值對照
LONG_TOKENS = {"long", "buy", "b", "做多", "多", "買", "買進", "1"}
SHORT_TOKENS = {"short", "sell", "s", "做空", "空", "賣", "賣出", "放空", "-1"}

# market 推斷:代號特徵
_TW_PATTERN = re.compile(r"^\d{4,6}[A-Z]?$")        # 台股多為 4~6 碼數字(如 2330、00878)
_CRYPTO_PATTERN = re.compile(r"(USDT|USDC|BTC|ETH|USD)$", re.IGNORECASE)
_US_PATTERN = re.compile(r"^[A-Z]{1,5}$")            # 美股多為 1~5 個大寫字母


def _norm(name: str) -> str:
    return re.sub(r"\s+", "", str(name)).strip().lower()


def _build_field_map(columns: Iterable[str]) -> dict[str, str]:
    """把實際欄位名對應到標準欄位名。回傳 {標準名: 實際欄位名}。"""
    normalized = {_norm(c): c for c in columns}
    field_map: dict[str, str] = {}
    for std, synonyms in FIELD_SYNONYMS.items():
        for syn in synonyms:
            key = _norm(syn)
            if key in normalized:
                field_map[std] = normalized[key]
                break
    return field_map


def infer_market(symbol: str, hint: Market | None = None) -> Market:
    """從標的代號推斷市場別。使用者有指定 hint 時優先採用。"""
    if hint and hint != Market.UNKNOWN:
        return hint
    s = str(symbol).strip().upper()
    if _CRYPTO_PATTERN.search(s):
        return Market.CRYPTO
    if _TW_PATTERN.match(s):
        return Market.TW_STOCK
    if _US_PATTERN.match(s):
        return Market.US_STOCK
    return Market.UNKNOWN


def _parse_side(value: Any) -> Side:
    v = _norm(value)
    if v in SHORT_TOKENS:
        return Side.SHORT
    if v in LONG_TOKENS:
        return Side.LONG
    # 預設視為做多(最常見);無法判斷時不該悄悄出錯,但也不該卡住分析
    return Side.LONG


def _parse_time(value: Any) -> datetime:
    if isinstance(value, datetime):
        return value
    s = str(value).strip()
    # 嘗試多種常見格式
    formats = [
        "%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d",
        "%Y%m%d", "%m/%d/%Y", "%d/%m/%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    # 最後嘗試 ISO 格式(含時區)
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except ValueError as exc:
        raise ValueError(f"無法解析時間格式: {value!r}") from exc


def _to_float(value: Any, default: float | None = None) -> float | None:
    if value is None or value == "":
        return default
    if isinstance(value, (int, float)):
        return float(value)
    # 去除千分位逗號、貨幣符號、空白
    s = re.sub(r"[,$￥¥\s]", "", str(value))
    if s in ("", "-", "—"):
        return default
    try:
        return float(s)
    except ValueError:
        return default


def sniff_format(path: str | Path) -> str:
    """依副檔名判斷格式: csv / json / excel。"""
    ext = Path(path).suffix.lower()
    if ext in (".csv", ".txt", ".tsv"):
        return "csv"
    if ext == ".json":
        return "json"
    if ext in (".xlsx", ".xls"):
        return "excel"
    raise ValueError(f"不支援的副檔名: {ext}(支援 .csv / .json / .xlsx)")


def _rows_from_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    import csv

    delimiter = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        rows = list(reader)
        columns = reader.fieldnames or []
    return list(columns), rows


def _rows_from_json(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    # 支援 [{...}, ...] 或 {"trades": [...]}
    if isinstance(data, dict):
        for key in ("trades", "data", "records", "orders"):
            if key in data and isinstance(data[key], list):
                data = data[key]
                break
        else:
            raise ValueError("JSON 物件中找不到交易陣列(預期鍵: trades/data/records)")
    if not isinstance(data, list) or not data:
        raise ValueError("JSON 必須是非空的交易陣列")
    columns = list(data[0].keys())
    return columns, data


def _rows_from_excel(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError(
            "讀取 Excel 需要 openpyxl。請執行: pip install openpyxl"
        ) from exc
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header = [str(c) if c is not None else "" for c in next(rows_iter)]
    rows: list[dict[str, Any]] = []
    for raw in rows_iter:
        if all(c is None for c in raw):
            continue
        rows.append(dict(zip(header, raw)))
    wb.close()
    return header, rows


def load_trades(
    path: str | Path,
    *,
    market_hint: Market | None = None,
    auto_estimate_costs: bool = True,
    field_overrides: Optional[dict[str, str]] = None,
) -> TradeLog:
    """從檔案載入交易紀錄,回傳正規化的 TradeLog。

    Args:
        path:               資料檔路徑(.csv / .json / .xlsx)
        market_hint:        若所有交易都屬同一市場,可在此指定以提升準確度
        auto_estimate_costs: 當該筆交易未提供 fees 時,是否用成本模型自動估算。
                             強烈建議開啟 — 忽略成本是賭徒最常見的自我欺騙。
        field_overrides:    手動指定欄位對應 {標準名: 實際欄位名},覆寫自動辨識

    Raises:
        ValueError: 缺少必要欄位,或資料無法解析
    """
    path = Path(path)
    fmt = sniff_format(path)

    if fmt == "csv":
        columns, rows = _rows_from_csv(path)
    elif fmt == "json":
        columns, rows = _rows_from_json(path)
    else:
        columns, rows = _rows_from_excel(path)

    if not rows:
        raise ValueError(f"檔案中沒有任何資料列: {path}")

    field_map = _build_field_map(columns)
    if field_overrides:
        field_map.update(field_overrides)

    # 檢查必要欄位。pnl 可由價格推算,故 pnl 與(進出場價)二擇一即可。
    required_core = ["symbol", "entry_price", "exit_price", "quantity"]
    has_prices = all(k in field_map for k in ("entry_price", "exit_price", "quantity"))
    has_pnl = "pnl" in field_map and "symbol" in field_map

    if not (has_prices or has_pnl):
        missing = [k for k in required_core if k not in field_map]
        raise ValueError(
            f"無法辨識必要欄位: {missing}。\n"
            f"偵測到的欄位: {columns}\n"
            f"請用 field_overrides 手動指定,例如 "
            f"{{'symbol': '你的代號欄位', 'entry_price': '你的買價欄位'}}"
        )

    def get(row: dict, std: str, default: Any = None) -> Any:
        col = field_map.get(std)
        return row.get(col, default) if col else default

    # 台股「張」單位偵測:若數量欄名含「張」(如 張數、張),代表單位是「張」
    # 而非「股」,1 張 = 1000 股。不換算會讓損益與成本差 1000 倍且不報錯,
    # 對台股使用者是最危險的靜默錯誤。
    qty_col = field_map.get("quantity", "")
    qty_in_lots = ("張" in str(qty_col))
    lot_multiplier = 1000 if qty_in_lots else 1

    trades: list[Trade] = []
    skipped = 0
    skip_reasons: list[str] = []   # 收集略過原因,回報給使用者(不再靜默吞錯)
    for row in rows:
        row_no = len(trades) + skipped + 1
        symbol = str(get(row, "symbol", "")).strip()
        if not symbol:
            skipped += 1
            if len(skip_reasons) < 10:
                skip_reasons.append(f"第 {row_no} 列:缺少標的代號")
            continue

        # 時間解析失敗是「這列資料髒」的明確訊號,單獨捕捉並記錄原因。
        try:
            entry_time = _parse_time(get(row, "entry_time", "1970-01-01"))
            exit_time = (
                _parse_time(get(row, "exit_time")) if get(row, "exit_time") else entry_time
            )
        except ValueError as exc:
            skipped += 1
            if len(skip_reasons) < 10:
                skip_reasons.append(f"第 {row_no} 列({symbol}):時間格式無法解析 — {exc}")
            continue

        market = infer_market(symbol, market_hint)
        side = _parse_side(get(row, "side", "long"))

        entry_price = _to_float(get(row, "entry_price"), 0.0) or 0.0
        exit_price = _to_float(get(row, "exit_price"), 0.0) or 0.0
        quantity = (_to_float(get(row, "quantity"), 0.0) or 0.0) * lot_multiplier

        fees = _to_float(get(row, "fees"), None)
        pnl = _to_float(get(row, "pnl"), None)
        tag = get(row, "tag")
        tag = str(tag).strip() if tag not in (None, "") else None

        # 成本處理:使用者沒給 fees 且開啟自動估算時,補上估計成本
        if fees is None and auto_estimate_costs and entry_price and quantity:
            # 當沖判定:進出場為同一日曆日(台股當沖證交稅減半)
            is_day_trade = entry_time.date() == exit_time.date()
            fees = estimate_round_trip_cost(
                market, side, entry_price, exit_price, quantity,
                is_day_trade=is_day_trade,
            )
        fees = fees or 0.0

        trade = Trade(
            symbol=symbol,
            market=market,
            side=side,
            entry_time=entry_time,
            exit_time=exit_time,
            entry_price=entry_price,
            exit_price=exit_price,
            quantity=quantity,
            fees=fees,
            pnl=pnl,  # 若為 None,Trade.__post_init__ 會用價格推算(已含 fees)
            tag=tag,
        )
        trades.append(trade)

    if not trades:
        detail = "\n  ".join(skip_reasons) if skip_reasons else "請檢查欄位對應與資料格式。"
        raise ValueError(
            f"沒有任何有效交易可解析(略過 {skipped} 列)。\n  {detail}"
        )

    # 略過比例過高,通常代表欄位對應整個錯了 — 主動警告而非靜默。
    skip_ratio = skipped / (len(trades) + skipped) if (len(trades) + skipped) else 0
    warn = ""
    if skip_ratio > 0.2:
        warn = f", ⚠️略過比例 {skip_ratio:.0%} 偏高(可能欄位對應有誤)"
    lot_note = ", 數量以『張』×1000 換算為股" if qty_in_lots else ""

    return TradeLog(
        trades=trades,
        source=f"{path.name} ({fmt}, 載入 {len(trades)} 筆, 略過 {skipped} 筆{warn}{lot_note})",
        account_label=path.stem,
    )
